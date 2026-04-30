"""
PDF va Excel eksport: bemor tarixi, hisobotlar.
"""
from io import BytesIO

from django.db.models import Count, Sum
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table as RLTable, TableStyle,
)
from reportlab.lib import colors

from apps.patients.models import Patient
from apps.visits.models import Visit


def patient_history_pdf(patient_id: int) -> bytes:
    """Bemor tarixi PDF formatida."""
    patient = Patient.objects.select_related('discount_category').get(pk=patient_id)
    visits = (
        Visit.objects
        .filter(patient=patient)
        .select_related('primary_doctor', 'primary_doctor__specialty')
        .order_by('-visit_date')
    )

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        rightMargin=2*cm, leftMargin=2*cm,
        topMargin=2*cm, bottomMargin=2*cm,
    )
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph(f"<b>Bemor tarixi:</b> {patient.full_name}", styles['Title']))
    elements.append(Spacer(1, 0.5*cm))
    elements.append(Paragraph(
        f"<b>Telefon:</b> {patient.phone}<br/>"
        f"<b>Tug'ilgan:</b> {patient.birth_date} ({patient.age} yosh)<br/>"
        f"<b>Manzil:</b> {patient.address or '—'}<br/>"
        f"<b>Chegirma:</b> {patient.discount_category.name if patient.discount_category else '—'}",
        styles['Normal']
    ))
    elements.append(Spacer(1, 0.5*cm))

    if not visits.exists():
        elements.append(Paragraph("Murojaatlar yo'q.", styles['Italic']))
    else:
        data = [['Sana', 'Shifokor', 'Mutaxassislik', 'Tashxis', 'Summa', 'Holat']]
        for v in visits:
            data.append([
                v.visit_date.strftime('%Y-%m-%d'),
                v.primary_doctor.full_name,
                v.primary_doctor.specialty.name,
                (v.diagnosis or '—')[:40],
                f'{v.total_cost:,.0f}',
                v.get_payment_status_display(),
            ])

        table = RLTable(data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563eb')),
            ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
            ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID',       (0, 0), (-1, -1), 0.5, colors.grey),
            ('ALIGN',      (4, 1), (4, -1), 'RIGHT'),
            ('FONTSIZE',   (0, 0), (-1, -1), 9),
        ]))
        elements.append(table)

    doc.build(elements)
    return buffer.getvalue()


def revenue_excel(start_date, end_date) -> bytes:
    """Berilgan davr daromadini Excel formatda."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Daromad hisoboti'

    headers = ['Sana', 'Murojaat soni', 'Brutto', 'Chegirma', 'Netto']
    ws.append(headers)

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='2563EB')
        cell.alignment = Alignment(horizontal='center')

    rows = (
        Visit.objects
        .filter(payment_status='PAID', visit_date__gte=start_date, visit_date__lte=end_date)
        .values('visit_date')
        .annotate(
            count=Count('id'),
            gross=Sum('subtotal'),
            discount=Sum('discount_amount'),
            net=Sum('total_cost'),
        )
        .order_by('visit_date')
    )

    for r in rows:
        ws.append([
            r['visit_date'].strftime('%Y-%m-%d'),
            r['count'],
            float(r['gross'] or 0),
            float(r['discount'] or 0),
            float(r['net'] or 0),
        ])

    for col in 'ABCDE':
        ws.column_dimensions[col].width = 18

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
